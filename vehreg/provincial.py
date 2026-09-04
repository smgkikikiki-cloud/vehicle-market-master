"""Provincial vehicle-registration warehouse and curated geographic analytics.

Provincial rows describe the same registration activity as the national DLT
files, only at a finer geographic grain. They therefore live in a separate fact
table and must never be unioned into ``fact_registration``: doing that would
double-count the market whenever a query did not explicitly filter province.

The raw provincial source may be very broad. Publication is a second decision:
``data/research/geo_publish.csv`` is an explicit editorial whitelist used by the
reader-facing Regional Market page. Admin/research code can still query every
matched provincial row.
"""

from __future__ import annotations

import csv
import hashlib
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Optional, Sequence

from .catalog import Catalog
from .db import register_source
from .ingest import Resolver
from .taxonomy import Grain

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_PUBLICATION_FILE = ROOT / "data" / "research" / "geo_publish.csv"
DEFAULT_REGISTRATION_TYPES = ("RY1", "RY2", "RY3")

SCHEMA = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS fact_registration_province (
    provincial_fact_id INTEGER PRIMARY KEY AUTOINCREMENT,
    period              TEXT NOT NULL,
    registration_type   TEXT NOT NULL,
    province            TEXT NOT NULL,
    model_id            TEXT NOT NULL,
    units               REAL NOT NULL,
    source_id           INTEGER NOT NULL REFERENCES dim_source(source_id),
    raw_brand           TEXT,
    raw_model           TEXT,
    raw_label           TEXT,
    match_how           TEXT,
    match_score         REAL,
    UNIQUE (
        period, registration_type, province, model_id, source_id,
        raw_brand, raw_model
    )
);
CREATE INDEX IF NOT EXISTS ix_provincial_period
    ON fact_registration_province(period);
CREATE INDEX IF NOT EXISTS ix_provincial_model_period
    ON fact_registration_province(model_id, period);
CREATE INDEX IF NOT EXISTS ix_provincial_province_period
    ON fact_registration_province(province, period);

CREATE TABLE IF NOT EXISTS provincial_review (
    review_id           INTEGER PRIMARY KEY AUTOINCREMENT,
    source_id           INTEGER NOT NULL REFERENCES dim_source(source_id),
    period              TEXT,
    registration_type   TEXT,
    province            TEXT,
    raw_brand           TEXT,
    raw_model           TEXT,
    units               REAL,
    reason              TEXT NOT NULL,
    best_guess          TEXT,
    score               REAL,
    status              TEXT NOT NULL DEFAULT 'open'
);
CREATE INDEX IF NOT EXISTS ix_provincial_review_period
    ON provincial_review(period, status);
"""

VIEW_SQL = """
DROP VIEW IF EXISTS provincial_classified;
CREATE VIEW provincial_classified AS
SELECT f.provincial_fact_id, f.period,
       f.registration_type AS fact_registration_type,
       f.province, f.model_id AS unit_id, 'MODEL' AS grain,
       f.units, f.source_id, f.raw_brand, f.raw_model, f.raw_label,
       f.match_how, f.match_score,
       d.catalog_year, d.brand, d.nameplate, d.model, d.segment,
       d.body_type, d.cab_type, d.market_position, d.powertrain,
       d.powertrain_group, d.origin_country, d.import_type,
       d.brand_segment, d.oem_group, d.brand_origin, d.drivetrain,
       d.market_scope, d.price_thb, d.price_min_thb, d.price_max_thb,
       d.battery_kwh
FROM fact_registration_province f
LEFT JOIN dim_unit d
  ON d.unit_id = f.model_id
 AND d.grain = 'MODEL'
 AND d.catalog_year = CAST(substr(f.period, 1, 4) AS INTEGER);
"""

THAI_MONTHS = {
    "มกราคม": 1,
    "กุมภาพันธ์": 2,
    "มีนาคม": 3,
    "เมษายน": 4,
    "พฤษภาคม": 5,
    "มิถุนายน": 6,
    "กรกฎาคม": 7,
    "สิงหาคม": 8,
    "กันยายน": 9,
    "ตุลาคม": 10,
    "พฤศจิกายน": 11,
    "ธันวาคม": 12,
    "ม.ค.": 1,
    "ก.พ.": 2,
    "มี.ค.": 3,
    "เม.ย.": 4,
    "พ.ค.": 5,
    "มิ.ย.": 6,
    "ก.ค.": 7,
    "ส.ค.": 8,
    "ก.ย.": 9,
    "ต.ค.": 10,
    "พ.ย.": 11,
    "ธ.ค.": 12,
}

# Region buckets are a TDR reporting convention, not a DLT legal field.
PROVINCE_REGION = {
    "กรุงเทพมหานคร": "BANGKOK",
    "นนทบุรี": "CENTRAL", "ปทุมธานี": "CENTRAL",
    "สมุทรปราการ": "CENTRAL", "นครปฐม": "CENTRAL",
    "สมุทรสาคร": "CENTRAL", "สมุทรสงคราม": "CENTRAL",
    "พระนครศรีอยุธยา": "CENTRAL", "อ่างทอง": "CENTRAL",
    "ลพบุรี": "CENTRAL", "สิงห์บุรี": "CENTRAL",
    "ชัยนาท": "CENTRAL", "สระบุรี": "CENTRAL", "สุพรรณบุรี": "CENTRAL",
    "ชลบุรี": "EAST", "ระยอง": "EAST", "จันทบุรี": "EAST",
    "ตราด": "EAST", "ฉะเชิงเทรา": "EAST", "ปราจีนบุรี": "EAST",
    "นครนายก": "EAST", "สระแก้ว": "EAST",
    "กาญจนบุรี": "WEST", "ราชบุรี": "WEST", "เพชรบุรี": "WEST",
    "ประจวบคีรีขันธ์": "WEST",
    "เชียงใหม่": "NORTH", "เชียงราย": "NORTH", "ลำปาง": "NORTH",
    "ลำพูน": "NORTH", "แม่ฮ่องสอน": "NORTH", "น่าน": "NORTH",
    "พะเยา": "NORTH", "แพร่": "NORTH", "อุตรดิตถ์": "NORTH",
    "พิษณุโลก": "NORTH", "สุโขทัย": "NORTH", "เพชรบูรณ์": "NORTH",
    "พิจิตร": "NORTH", "กำแพงเพชร": "NORTH", "นครสวรรค์": "NORTH",
    "อุทัยธานี": "NORTH", "ตาก": "NORTH",
    "นครราชสีมา": "NORTHEAST", "บุรีรัมย์": "NORTHEAST",
    "สุรินทร์": "NORTHEAST", "ศรีสะเกษ": "NORTHEAST",
    "อุบลราชธานี": "NORTHEAST", "ยโสธร": "NORTHEAST",
    "ชัยภูมิ": "NORTHEAST", "อำนาจเจริญ": "NORTHEAST",
    "บึงกาฬ": "NORTHEAST", "หนองบัวลำภู": "NORTHEAST",
    "ขอนแก่น": "NORTHEAST", "อุดรธานี": "NORTHEAST",
    "เลย": "NORTHEAST", "หนองคาย": "NORTHEAST",
    "มหาสารคาม": "NORTHEAST", "ร้อยเอ็ด": "NORTHEAST",
    "กาฬสินธุ์": "NORTHEAST", "สกลนคร": "NORTHEAST",
    "นครพนม": "NORTHEAST", "มุกดาหาร": "NORTHEAST",
    "ชุมพร": "SOUTH", "ระนอง": "SOUTH", "สุราษฎร์ธานี": "SOUTH",
    "พังงา": "SOUTH", "ภูเก็ต": "SOUTH", "กระบี่": "SOUTH",
    "นครศรีธรรมราช": "SOUTH", "ตรัง": "SOUTH", "พัทลุง": "SOUTH",
    "สตูล": "SOUTH", "สงขลา": "SOUTH", "ปัตตานี": "SOUTH",
    "ยะลา": "SOUTH", "นราธิวาส": "SOUTH",
}

PROVINCE_ALIASES = {
    "กรุงเทพฯ": "กรุงเทพมหานคร",
    "กทม": "กรุงเทพมหานคร",
    "กทม.": "กรุงเทพมหานคร",
    "อยุธยา": "พระนครศรีอยุธยา",
    "ประจวบฯ": "ประจวบคีรีขันธ์",
}

_REGISTRATION_RE = re.compile(r"(?:รย\.?|RY)\s*([0-9]+)", re.IGNORECASE)


@dataclass(frozen=True)
class PublishedModelRule:
    category: str
    category_label: str
    label: str
    brand: str
    model_pattern: str
    display_order: int = 999
    enabled: bool = True
    note: str = ""


@dataclass
class ProvincialIngestReport:
    source_id: int
    rows_read: int = 0
    facts_written: int = 0
    review_rows: int = 0
    units_ingested: float = 0.0
    units_review: float = 0.0
    units_skipped_class: float = 0.0
    units_skipped_no_catalog: float = 0.0
    unchanged: bool = False

    @property
    def units_seen(self) -> float:
        return (
            self.units_ingested
            + self.units_review
            + self.units_skipped_class
            + self.units_skipped_no_catalog
        )

    @property
    def match_coverage(self) -> float:
        denominator = self.units_ingested + self.units_review
        return self.units_ingested / denominator if denominator else 0.0

    def render(self) -> str:
        if self.unchanged:
            return "provincial source unchanged; existing facts kept"
        return "\n".join(
            [
                f"rows read             : {self.rows_read:,}",
                f"facts written         : {self.facts_written:,}",
                f"review rows           : {self.review_rows:,}",
                f"units ingested        : {self.units_ingested:,.0f}",
                f"units in review       : {self.units_review:,.0f}",
                f"units skipped class   : {self.units_skipped_class:,.0f}",
                f"units no catalog year : {self.units_skipped_no_catalog:,.0f}",
                f"match coverage        : {self.match_coverage:.1%}",
            ]
        )


def ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(SCHEMA)
    conn.executescript(VIEW_SQL)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


def _text(value: object) -> str:
    return str(value or "").strip()


def _units(value: object) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def period_from_thai(year: object, month: object) -> Optional[str]:
    try:
        y = int(float(str(year).strip()))
    except (TypeError, ValueError):
        return None
    if y >= 2400:
        y -= 543
    month_text = _text(month)
    m = THAI_MONTHS.get(month_text)
    if m is None:
        try:
            m = int(float(month_text))
        except (TypeError, ValueError):
            return None
    if not 1 <= m <= 12:
        return None
    return f"{y:04d}-{m:02d}"


def registration_type(raw: object) -> Optional[str]:
    text = _text(raw)
    match = _REGISTRATION_RE.search(text)
    return f"RY{int(match.group(1))}" if match else None


def normalize_province(raw: object) -> str:
    text = " ".join(_text(raw).replace("จังหวัด", "").split())
    return PROVINCE_ALIASES.get(text, text)


def region_for(province: str) -> str:
    return PROVINCE_REGION.get(normalize_province(province), "UNKNOWN")


def load_publication_rules(
    path: Path | str = DEFAULT_PUBLICATION_FILE,
    *,
    enabled_only: bool = True,
) -> list[PublishedModelRule]:
    source = Path(path)
    if not source.exists():
        return []
    out: list[PublishedModelRule] = []
    with source.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            enabled = _text(row.get("enabled", "1")).lower() not in {
                "0", "false", "no", "off"
            }
            rule = PublishedModelRule(
                category=_text(row.get("category")),
                category_label=_text(row.get("category_label")),
                label=_text(row.get("label")),
                brand=_text(row.get("brand")),
                model_pattern=_text(row.get("model_pattern")) or "%",
                display_order=int(_text(row.get("display_order")) or 999),
                enabled=enabled,
                note=_text(row.get("note")),
            )
            if rule.category and rule.label and (rule.enabled or not enabled_only):
                out.append(rule)
    return sorted(out, key=lambda r: (r.category, r.display_order, r.label))


def _upsert_source(conn: sqlite3.Connection, path: Path, digest: str,
                   source_name: str, notes: str) -> tuple[int, bool]:
    row = conn.execute(
        "SELECT source_id, file_sha256 FROM dim_source WHERE name=?",
        (source_name,),
    ).fetchone()
    if row and row["file_sha256"] == digest:
        existing = conn.execute(
            "SELECT 1 FROM fact_registration_province WHERE source_id=? LIMIT 1",
            (row["source_id"],),
        ).fetchone()
        if existing:
            return int(row["source_id"]), True

    if row:
        source_id = int(row["source_id"])
        with conn:
            conn.execute(
                "UPDATE dim_source SET publisher='DLT', file_name=?, file_sha256=?, "
                "notes=? WHERE source_id=?",
                (str(path), digest, notes, source_id),
            )
            conn.execute(
                "DELETE FROM fact_registration_province WHERE source_id=?",
                (source_id,),
            )
            conn.execute("DELETE FROM provincial_review WHERE source_id=?", (source_id,))
        return source_id, False

    source_id = register_source(
        conn,
        source_name,
        publisher="DLT",
        file_name=str(path),
        file_sha256=digest,
        notes=notes,
    )
    return source_id, False


def ingest_provincial_xlsx(
    conn: sqlite3.Connection,
    catalogs: Mapping[int, Catalog],
    path: Path | str,
    *,
    source_name: Optional[str] = None,
    sheet_name: str = "Data",
    registration_types: Sequence[str] = DEFAULT_REGISTRATION_TYPES,
    batch_size: int = 5000,
) -> ProvincialIngestReport:
    """Ingest the DLT ``brand × model × province`` workbook.

    Expected columns are the seven fields in the source's Data sheet:
    ปี, เดือน, ประเภทรถ, จังหวัด, ยี่ห้อรถ, รุ่นรถ, จำนวนรถ.

    Facts are deliberately folded to canonical MODEL grain. Provincial
    geography is a market-location layer; trim detail remains in the separate
    national trim ledger.
    """
    ensure_schema(conn)
    source_path = Path(path)
    digest = _sha256(source_path)
    source_name = source_name or f"DLT Provincial | {source_path.name}"
    notes = (
        "DLT new registrations by year/month/vehicle class/province/brand/model. "
        "Province is registration location, not dealer retail-sale territory."
    )
    source_id, unchanged = _upsert_source(
        conn, source_path, digest, source_name, notes
    )
    report = ProvincialIngestReport(source_id=source_id, unchanged=unchanged)
    if unchanged:
        return report

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment guard
        raise RuntimeError(
            "Provincial XLSX import requires openpyxl; install requirements.txt"
        ) from exc

    allowed = {str(value).upper() for value in registration_types}
    resolvers = {year: Resolver(catalog, conn) for year, catalog in catalogs.items()}
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"{source_path}: sheet {sheet_name!r} not found; "
            f"available: {', '.join(workbook.sheetnames)}"
        )
    sheet = workbook[sheet_name]
    header = [
        _text(value)
        for value in next(
            sheet.iter_rows(min_row=6, max_row=6, min_col=1, max_col=7,
                            values_only=True)
        )
    ]
    expected = ["ปี", "เดือน", "ประเภทรถ", "จังหวัด", "ยี่ห้อรถ", "รุ่นรถ", "จำนวนรถ"]
    if header != expected:
        raise ValueError(
            f"{source_path}: unexpected Data header {header!r}; expected {expected!r}"
        )

    fact_batch: list[tuple[object, ...]] = []
    review_batch: list[tuple[object, ...]] = []

    def flush() -> None:
        if fact_batch:
            conn.executemany(
                "INSERT INTO fact_registration_province "
                "(period,registration_type,province,model_id,units,source_id,"
                "raw_brand,raw_model,raw_label,match_how,match_score) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                fact_batch,
            )
            fact_batch.clear()
        if review_batch:
            conn.executemany(
                "INSERT INTO provincial_review "
                "(source_id,period,registration_type,province,raw_brand,raw_model,"
                "units,reason,best_guess,score) VALUES (?,?,?,?,?,?,?,?,?,?)",
                review_batch,
            )
            review_batch.clear()
        conn.commit()

    for values in sheet.iter_rows(min_row=7, min_col=1, max_col=7, values_only=True):
        report.rows_read += 1
        year_raw, month_raw, type_raw, province_raw, brand_raw, model_raw, units_raw = values
        units = _units(units_raw)
        if units is None or units == 0:
            continue
        period = period_from_thai(year_raw, month_raw)
        reg = registration_type(type_raw)
        if reg not in allowed:
            report.units_skipped_class += units
            continue
        if not period:
            review_batch.append(
                (source_id, None, reg, normalize_province(province_raw),
                 _text(brand_raw), _text(model_raw), units,
                 "period-not-readable", None, None)
            )
            report.review_rows += 1
            report.units_review += units
            continue
        year = int(period[:4])
        catalog = catalogs.get(year)
        resolver = resolvers.get(year)
        if catalog is None or resolver is None:
            report.units_skipped_no_catalog += units
            continue

        province = normalize_province(province_raw)
        raw_brand = _text(brand_raw)
        raw_model = _text(model_raw)
        if not province or not raw_brand or not raw_model:
            review_batch.append(
                (source_id, period, reg, province, raw_brand, raw_model, units,
                 "missing-province-brand-or-model", None, None)
            )
            report.review_rows += 1
            report.units_review += units
            continue

        unit_id, grain, how, score, reason = resolver.resolve(
            raw_brand, raw_model, reg=reg
        )
        if unit_id is None or grain is Grain.BRAND:
            review_batch.append(
                (source_id, period, reg, province, raw_brand, raw_model, units,
                 reason or "model-not-resolved", unit_id, score)
            )
            report.review_rows += 1
            report.units_review += units
        else:
            model_id = (
                catalog.model_for_variant(unit_id).id
                if grain is Grain.VARIANT
                else unit_id
            )
            if model_id not in catalog.models:
                review_batch.append(
                    (source_id, period, reg, province, raw_brand, raw_model, units,
                     "resolved-unit-is-not-model", model_id, score)
                )
                report.review_rows += 1
                report.units_review += units
            else:
                fact_batch.append(
                    (
                        period, reg, province, model_id, units, source_id,
                        raw_brand, raw_model, f"{raw_brand} {raw_model}".strip(),
                        how, score,
                    )
                )
                report.facts_written += 1
                report.units_ingested += units

        if len(fact_batch) + len(review_batch) >= batch_size:
            flush()

    flush()
    workbook.close()
    ensure_schema(conn)
    return report


def available_periods(conn: sqlite3.Connection) -> list[str]:
    ensure_schema(conn)
    return [
        str(row["period"])
        for row in conn.execute(
            "SELECT DISTINCT period FROM fact_registration_province ORDER BY period"
        )
    ]


def _window_where(period_from: str, period_to: str,
                  registration_types: Sequence[str]) -> tuple[str, list[object]]:
    regs = [str(value).upper() for value in registration_types]
    placeholders = ",".join("?" for _ in regs)
    return (
        f"f.period BETWEEN ? AND ? AND f.registration_type IN ({placeholders})",
        [period_from, period_to, *regs],
    )


def _rule_province_units(
    conn: sqlite3.Connection,
    rule: PublishedModelRule,
    period_from: str,
    period_to: str,
    registration_types: Sequence[str] = DEFAULT_REGISTRATION_TYPES,
) -> dict[str, float]:
    where, params = _window_where(period_from, period_to, registration_types)
    rows = conn.execute(
        "SELECT f.province, SUM(f.units) AS units "
        "FROM fact_registration_province f "
        "JOIN dim_unit d ON d.unit_id=f.model_id AND d.grain='MODEL' "
        "AND d.catalog_year=CAST(substr(f.period,1,4) AS INTEGER) "
        f"WHERE {where} AND d.brand=? AND d.model LIKE ? "
        "GROUP BY f.province",
        [*params, rule.brand, rule.model_pattern],
    ).fetchall()
    return {str(row["province"]): float(row["units"] or 0) for row in rows}


def geographic_profile(
    conn: sqlite3.Connection,
    rule: PublishedModelRule,
    category_rules: Iterable[PublishedModelRule],
    period_from: str,
    period_to: str,
    registration_types: Sequence[str] = DEFAULT_REGISTRATION_TYPES,
) -> list[dict[str, object]]:
    """Province footprint + within-category share and geographic over-index.

    ``over_index`` compares the selected model's share of its curated category
    in a province with its share of that category nationwide. A value of 1.5
    means the model is 50% more concentrated there than its national competitive
    position would imply. It is not a dealer-sales metric.
    """
    ensure_schema(conn)
    selected = _rule_province_units(
        conn, rule, period_from, period_to, registration_types
    )
    category: dict[str, float] = defaultdict(float)
    for candidate in category_rules:
        for province, units in _rule_province_units(
            conn, candidate, period_from, period_to, registration_types
        ).items():
            category[province] += units

    selected_total = sum(selected.values())
    category_total = sum(category.values())
    national_category_share = (
        selected_total / category_total if category_total else 0.0
    )
    provinces = sorted(set(category) | set(selected))
    rows: list[dict[str, object]] = []
    for province in provinces:
        units = selected.get(province, 0.0)
        category_units = category.get(province, 0.0)
        local_share = units / category_units if category_units else 0.0
        over_index = (
            local_share / national_category_share
            if national_category_share else None
        )
        rows.append(
            {
                "province": province,
                "region": region_for(province),
                "units": units,
                "distribution_share": units / selected_total if selected_total else 0.0,
                "category_units": category_units,
                "local_category_share": local_share,
                "national_category_share": national_category_share,
                "over_index": over_index,
            }
        )
    return sorted(rows, key=lambda row: (-float(row["units"]), str(row["province"])))


def regional_profile(rows: Iterable[Mapping[str, object]]) -> list[dict[str, object]]:
    selected: dict[str, float] = defaultdict(float)
    category: dict[str, float] = defaultdict(float)
    for row in rows:
        region = str(row.get("region") or "UNKNOWN")
        selected[region] += float(row.get("units") or 0)
        category[region] += float(row.get("category_units") or 0)
    total = sum(selected.values())
    category_total = sum(category.values())
    national_share = total / category_total if category_total else 0.0
    out: list[dict[str, object]] = []
    for region in sorted(category):
        units = selected[region]
        cat_units = category[region]
        local_share = units / cat_units if cat_units else 0.0
        out.append(
            {
                "region": region,
                "units": units,
                "distribution_share": units / total if total else 0.0,
                "local_category_share": local_share,
                "over_index": local_share / national_share if national_share else None,
            }
        )
    return sorted(out, key=lambda row: -float(row["units"]))


def category_competition(
    conn: sqlite3.Connection,
    rules: Iterable[PublishedModelRule],
    period_from: str,
    period_to: str,
    *,
    province: Optional[str] = None,
    registration_types: Sequence[str] = DEFAULT_REGISTRATION_TYPES,
) -> list[dict[str, object]]:
    out: list[dict[str, object]] = []
    for rule in rules:
        by_province = _rule_province_units(
            conn, rule, period_from, period_to, registration_types
        )
        units = (
            by_province.get(normalize_province(province), 0.0)
            if province
            else sum(by_province.values())
        )
        out.append({"label": rule.label, "units": units})
    total = sum(float(row["units"]) for row in out)
    for row in out:
        row["share"] = float(row["units"]) / total if total else 0.0
    out.sort(key=lambda row: (-float(row["units"]), str(row["label"])))
    for rank, row in enumerate(out, 1):
        row["rank"] = rank
    return out


def reconciliation_for_period(conn: sqlite3.Connection, period: str) -> dict[str, float]:
    """Compare provincial sums with canonical national facts where both exist."""
    ensure_schema(conn)
    provincial = float(
        conn.execute(
            "SELECT COALESCE(SUM(units),0) AS units "
            "FROM fact_registration_province WHERE period=?",
            (period,),
        ).fetchone()["units"]
        or 0
    )
    national = float(
        conn.execute(
            "SELECT COALESCE(SUM(units),0) AS units FROM fact_registration "
            "WHERE period=? AND province='ALL' AND registration_type IN ('RY1','RY2','RY3')",
            (period,),
        ).fetchone()["units"]
        or 0
    )
    return {
        "period": period,
        "provincial_units": provincial,
        "national_units": national,
        "difference": provincial - national,
        "difference_pct": ((provincial - national) / national if national else 0.0),
    }
